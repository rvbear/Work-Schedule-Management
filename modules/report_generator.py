from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows


class ReportGenerator:
    """엑셀 리포트 생성을 담당하는 클래스"""
    
    @staticmethod
    def create_monthly_summary_report(daily_data, employee_info, output_path):
        """
        부서별 월 합산 리포트 생성
        
        Args:
            daily_data: 일별 근태 데이터 DataFrame
            employee_info: 사원 정보 DataFrame
            output_path: 출력 파일 경로
        """
        # 사원 정보의 문자열 컬럼 타입 확인 및 변환
        employee_info = employee_info.copy()
        employee_info['사원명'] = employee_info['사원명'].astype(str)
        employee_info['부서명'] = employee_info['부서명'].astype(str)
        
        # 사원 정보와 병합
        merged_data = daily_data.merge(
            employee_info[['카드번호', '사원명', '부서명', '부서코드']],
            on='카드번호',
            how='left'
        )
        
        # 부서별로 그룹화하여 월 합산
        summary = merged_data.groupby(['부서명', '사원명']).agg({
            'work_ot': 'sum',
            'overtime': 'sum',
            'basic_pay': 'sum',
            'approved_ot': 'sum',
            'night_work': 'sum',
            'holiday_bonus': 'sum',
            'meal_allowance': 'sum',
            'transport_allowance': 'sum'
        }).reset_index()
        
        # 컬럼명 변경
        summary.columns = ['부서명', '성명', '근무O/T', '연장', '기본급', '인정 OT', 
                          '야간적용', '휴일추가', '식대', '교통비']
        
        # 엑셀 파일 생성
        wb = Workbook()
        
        # 부서별로 시트 생성
        departments = summary['부서명'].unique()
        
        for idx, dept in enumerate(departments):
            # 부서명을 문자열로 변환하고 유효한 시트명 생성
            dept_name = str(dept).strip()[:31]
            
            if idx == 0:
                ws = wb.active
                ws.title = dept_name
            else:
                ws = wb.create_sheet(title=dept_name)
            
            # 해당 부서 데이터 필터링
            dept_data = summary[summary['부서명'] == dept].drop(columns=['부서명'])
            
            # 헤더 스타일 설정
            header_font = Font(bold=True, size=11)
            header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 데이터 작성
            for r_idx, row in enumerate(dataframe_to_rows(dept_data, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    if r_idx == 1:  # 헤더
                        cell.font = header_font
                        cell.fill = header_fill
            
            # 열 너비 자동 조정
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(output_path)
        print(f"월간 합산 리포트 생성 완료: {output_path}")
    
    
    @staticmethod
    def create_daily_detail_report(daily_data, employee_info, output_path):
        """
        개인별 일별 상세 리포트 생성
        
        Args:
            daily_data: 일별 근태 데이터 DataFrame
            employee_info: 사원 정보 DataFrame
            output_path: 출력 파일 경로
        """
        # 사원 정보의 문자열 컬럼 타입 확인 및 변환
        employee_info = employee_info.copy()
        employee_info['사원명'] = employee_info['사원명'].astype(str)
        employee_info['부서명'] = employee_info['부서명'].astype(str)
        
        # 사원 정보와 병합
        merged_data = daily_data.merge(
            employee_info[['카드번호', '사원명', '부서명']],
            on='카드번호',
            how='left'
        )
        
        # 정렬
        merged_data = merged_data.sort_values(['사원명', 'date'])
        
        # 필요한 컬럼만 선택 및 순서 변경
        report_data = merged_data[[
            '사원명', 'date', 'check_in', 'check_out', 'work_ot', 'overtime',
            'basic_pay', 'approved_ot', 'night_work', 'holiday_bonus',
            'meal_allowance', 'transport_allowance'
        ]].copy()
        
        # 컬럼명 변경
        report_data.columns = ['성명', '날짜', '출근', '퇴근', '근무 OT', '연장', 
                              '기본급', '인정 OT', '야간적용', '휴일추가', 
                              '식대', '교통비']
        
        # 엑셀 파일 생성
        wb = Workbook()
        
        # 사원별로 시트 생성
        employees = report_data['성명'].unique()
        
        for idx, emp in enumerate(employees):
            # 사원명을 문자열로 변환하고 유효한 시트명 생성
            emp_name = str(emp).strip()[:31]
            
            if idx == 0:
                ws = wb.active
                ws.title = emp_name
            else:
                ws = wb.create_sheet(title=emp_name)
            
            # 해당 사원 데이터 필터링
            emp_data = report_data[report_data['성명'] == emp]
            
            # 헤더 스타일 설정
            header_font = Font(bold=True, size=11)
            header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 데이터 작성
            for r_idx, row in enumerate(dataframe_to_rows(emp_data, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    if r_idx == 1:  # 헤더
                        cell.font = header_font
                        cell.fill = header_fill
            
            # 열 너비 자동 조정
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(output_path)
        print(f"일별 상세 리포트 생성 완료: {output_path}")
